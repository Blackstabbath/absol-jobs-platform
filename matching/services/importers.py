import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from matching.models import Candidate, CandidateExperience, CandidateProfile, CandidateSkill, Employer, Job


def _rows_from_csv(uploaded_file):
    uploaded_file.seek(0)
    decoded = uploaded_file.read().decode("utf-8-sig").splitlines()
    reader = csv.reader(decoded)
    try:
        headers = next(reader)
    except StopIteration:
        return []
    headers = unique_headers(headers)
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in reader
        if any(value not in (None, "") for value in row)
    ]


def _rows_from_xlsx(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = unique_headers([str(value or "").strip() for value in rows[0]])
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


def rows_from_upload(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(uploaded_file)
    if suffix in {".xlsx", ".xls"}:
        return _rows_from_xlsx(uploaded_file)
    raise ValueError("Upload must be a CSV or XLSX file.")


def unique_headers(headers):
    counts = {}
    unique = []
    for header in headers:
        clean = str(header or "").strip()
        key = clean or "unnamed"
        counts[key] = counts.get(key, 0) + 1
        unique.append(key if counts[key] == 1 else f"{key}__{counts[key]}")
    return unique


def normalise_label(value):
    value = re.sub(r"__\d+$", "", str(value or ""))
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def base_label(value):
    return re.sub(r"__\d+$", "", str(value or "")).strip()


def cell(row, *names, default=""):
    lower_map = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        value = lower_map.get(name.lower())
        if value not in (None, ""):
            return str(value).strip()
    return default


def fuzzy_cell(row, *patterns, default=""):
    for key, value in row.items():
        label = normalise_label(key)
        if value not in (None, "") and all(pattern.lower() in label for pattern in patterns):
            return str(value).strip()
    return default


def fuzzy_values(row, *patterns):
    values = []
    for key, value in row.items():
        label = normalise_label(key)
        if value not in (None, "") and all(pattern.lower() in label for pattern in patterns):
            values.append(str(value).strip())
    return values


def fuzzy_items(row, *patterns):
    items = []
    for key, value in row.items():
        label = normalise_label(key)
        if value not in (None, "") and all(pattern.lower() in label for pattern in patterns):
            items.append((key, str(value).strip()))
    return items


def first_available(row, exact_names=(), fuzzy_patterns=(), default=""):
    value = cell(row, *exact_names)
    if value:
        return value
    for patterns in fuzzy_patterns:
        value = fuzzy_cell(row, *patterns)
        if value:
            return value
    return default


def split_full_name(full_name):
    parts = str(full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def valid_candidate_status(value):
    allowed = {choice for choice, _label in Candidate.Status.choices}
    return value if value in allowed else Candidate.Status.ACTIVE


def bool_cell(row, *names):
    value = first_available(row, exact_names=names).lower()
    return value in {"1", "true", "yes", "y", "remote", "available"}


def decimal_cell(row, *names, default="0"):
    try:
        return Decimal(first_available(row, exact_names=names, default=default) or default)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def int_cell(row, *names):
    value = first_available(row, exact_names=names)
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def import_sarp_candidate(row, index):
    full_name = first_available(
        row,
        exact_names=("Full Name (as per UNHCR or ID)",),
        fuzzy_patterns=(("full name",),),
    )
    first_name = cell(row, "first_name", "first name", "firstname")
    last_name = cell(row, "last_name", "last name", "lastname")
    if full_name and not first_name:
        first_name, last_name = split_full_name(full_name)

    email = first_available(
        row,
        exact_names=("email", "candidate email", "Email Address"),
        fuzzy_patterns=(("email address",),),
    )
    if not email:
        return None, f"Row {index}: email is required."
    if not first_name:
        return None, f"Row {index}: full name or first_name is required."

    skill_items = fuzzy_items(row, "main field", "skill")
    skill_values = []
    skill_values.extend(value for _key, value in skill_items)
    skill_values.extend(fuzzy_values(row, "computer skill level"))
    skill_values = [value for value in skill_values if value and value.lower() not in {"none", "n/a", "na"}]

    year_items = fuzzy_items(row, "years", "work experience")
    detail_items = fuzzy_items(row, "details", "work experience")
    experience_values = [decimal_from_text(value) for _key, value in year_items]
    experience_values = [value for value in experience_values if value is not None]
    years_experience = max(experience_values) if experience_values else Decimal("0")

    work_details = [value for _key, value in detail_items]
    education_parts = [
        first_available(row, fuzzy_patterns=(("highest level", "education"),)),
        first_available(row, fuzzy_patterns=(("field of study",),)),
        first_available(row, fuzzy_patterns=(("university", "institution"),)),
        first_available(row, fuzzy_patterns=(("year of completion",),)),
        first_available(row, fuzzy_patterns=(("country where", "obtained"),)),
    ]
    certification_parts = [
        first_available(row, fuzzy_patterns=(("certificates", "diplomas", "licenses"),)),
        first_available(row, fuzzy_patterns=(("professional certifications",),)),
        first_available(row, fuzzy_patterns=(("educational certificates",),)),
    ]
    summary_parts = [
        labelled("Nationality", first_available(row, fuzzy_patterns=(("nationality",),))),
        labelled("SARP status", first_available(row, exact_names=("Status",))),
        labelled("Current location", first_available(row, fuzzy_patterns=(("currently living",),))),
        labelled("English speaking", first_available(row, fuzzy_patterns=(("speak english",),))),
        labelled("English reading/writing", first_available(row, fuzzy_patterns=(("read and write", "english"),))),
        labelled("Computer skills", first_available(row, fuzzy_patterns=(("basic computer skills",),))),
        labelled("Migration countries", first_available(row, fuzzy_patterns=(("countries", "interested", "migrating"),))),
        labelled("LinkedIn", first_available(row, fuzzy_patterns=(("linkedin",),))),
        labelled("Passport", first_available(row, fuzzy_patterns=(("valid passport",),))),
        labelled("Notes", first_available(row, fuzzy_patterns=(("anything else",),))),
    ]

    defaults = {
        "first_name": first_name,
        "last_name": last_name,
        "phone": first_available(row, exact_names=("phone", "WhatsApp Number (active contact)"), fuzzy_patterns=(("whatsapp",),)),
        "location": first_available(row, exact_names=("location",), fuzzy_patterns=(("currently living",),)),
        "willing_remote": True,
        "work_authorization": first_available(row, fuzzy_patterns=(("willing to migrate",), ("countries", "interested", "migrating"))),
        "years_experience": years_experience,
        "salary_expectation": None,
        "skills": ", ".join(dict.fromkeys(skill_values)),
        "education": "\n".join(part for part in education_parts if part),
        "certifications": "\n".join(part for part in certification_parts if part),
        "summary": "\n".join(part for part in summary_parts if part),
        "relevant_experience": "\n\n".join(work_details),
        "status": Candidate.Status.ACTIVE,
    }
    candidate, was_created = Candidate.objects.update_or_create(email=email, defaults=defaults)
    upsert_sarp_profile(candidate, row)
    upsert_sarp_skills_and_experience(candidate, skill_items, year_items, detail_items)
    return was_created, None


def upsert_sarp_profile(candidate, row):
    profile_defaults = {
        "raw_answers": {base_label(key): value for key, value in row.items()},
        "gender": first_available(row, fuzzy_patterns=(("gender",),)),
        "date_of_birth": first_available(row, fuzzy_patterns=(("date of birth",),)),
        "aims_number": first_available(row, fuzzy_patterns=(("aims number",),)),
        "nationality": first_available(row, fuzzy_patterns=(("nationality",),)),
        "arrival_year": first_available(row, fuzzy_patterns=(("year", "arrive"),)),
        "majlis": first_available(row, fuzzy_patterns=(("majlis",),)),
        "unhcr_registration": first_available(row, fuzzy_patterns=(("unhcr",), ("online registration",))),
        "marital_status": first_available(row, fuzzy_patterns=(("marital status",),)),
        "education_level": first_available(row, fuzzy_patterns=(("highest level", "education"),)),
        "field_of_study": first_available(row, fuzzy_patterns=(("field of study",),)),
        "institution": first_available(row, fuzzy_patterns=(("university",), ("institution name",))),
        "education_year": first_available(row, fuzzy_patterns=(("year of completion",),)),
        "education_country": first_available(row, fuzzy_patterns=(("country where", "obtained"),)),
        "education_verified": first_available(row, fuzzy_patterns=(("educational documents", "verified"),)),
        "english_speaking": first_available(row, fuzzy_patterns=(("speak english",),)),
        "english_reading_writing": first_available(row, fuzzy_patterns=(("read and write", "english"),)),
        "english_test": first_available(row, fuzzy_patterns=(("english language proficiency tests",),)),
        "english_score": first_available(row, fuzzy_patterns=(("what was your score",),)),
        "basic_computer_skills": first_available(row, fuzzy_patterns=(("basic computer skills",),)),
        "computer_skill_level": first_available(row, fuzzy_patterns=(("computer skill level",),)),
        "willing_to_migrate": first_available(row, fuzzy_patterns=(("willing to migrate",),)),
        "migration_countries": first_available(row, fuzzy_patterns=(("countries", "interested", "migrating"),)),
        "migration_reason": first_available(row, fuzzy_patterns=(("why did you choose",),)),
        "health_conditions": first_available(row, fuzzy_patterns=(("health conditions",),)),
        "linkedin_url": first_available(row, fuzzy_patterns=(("linkedin",),)),
        "valid_passport": first_available(row, fuzzy_patterns=(("valid passport",),)),
        "passport_expiry": first_available(row, fuzzy_patterns=(("expiry date",),)),
        "can_find_own_employer": first_available(row, fuzzy_patterns=(("find your own employer",),)),
        "source": first_available(row, fuzzy_patterns=(("how did you hear",),)),
        "consent_truthful": first_available(row, fuzzy_patterns=(("accurate", "truthful"),)),
        "consent_database": first_available(row, fuzzy_patterns=(("database purposes",),)),
        "consent_share": first_available(row, fuzzy_patterns=(("share my profile",),)),
        "consent_contact": first_available(row, fuzzy_patterns=(("contacted via",),)),
        "consent_more_documents": first_available(row, fuzzy_patterns=(("additional documentation",),)),
        "consent_data_storage": first_available(row, fuzzy_patterns=(("personal data", "stored"),)),
        "photo_document_upload": first_available(row, fuzzy_patterns=(("upload a photo",),)),
        "resume_upload": first_available(row, fuzzy_patterns=(("upload resume",),)),
        "education_upload": first_available(row, fuzzy_patterns=(("upload educational",),)),
        "certification_upload": first_available(row, fuzzy_patterns=(("upload professional",),)),
        "additional_notes": "\n".join(
            value
            for value in [
                first_available(row, fuzzy_patterns=(("anything else we should know",),)),
                first_available(row, fuzzy_patterns=(("skills experience or situation",),)),
            ]
            if value
        ),
    }
    CandidateProfile.objects.update_or_create(candidate=candidate, defaults=profile_defaults)


def upsert_sarp_skills_and_experience(candidate, skill_items, year_items, detail_items):
    CandidateSkill.objects.filter(candidate=candidate).delete()
    CandidateExperience.objects.filter(candidate=candidate).delete()
    max_len = max(len(skill_items), len(year_items), len(detail_items), 0)
    seen_skills = set()
    for index in range(max_len):
        skill = skill_items[index][1] if index < len(skill_items) else ""
        year_value = decimal_from_text(year_items[index][1]) if index < len(year_items) else None
        detail = detail_items[index][1] if index < len(detail_items) else ""
        source = base_label(skill_items[index][0]) if index < len(skill_items) else ""
        if skill:
            skill_key = skill.lower().strip()
            if skill_key not in seen_skills:
                CandidateSkill.objects.create(
                    candidate=candidate,
                    name=skill,
                    years_experience=year_value,
                    source_label=source,
                )
                seen_skills.add(skill_key)
        if skill or detail:
            CandidateExperience.objects.create(
                candidate=candidate,
                skill=skill,
                years_experience=year_value,
                details=detail,
                source_label=source,
            )


def decimal_from_text(value):
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    if not match:
        return None
    try:
        parsed = Decimal(match.group())
    except InvalidOperation:
        return None
    if parsed > Decimal("80"):
        return None
    return parsed


def labelled(label, value):
    return f"{label}: {value}" if value else ""


def import_candidates(uploaded_file):
    created = updated = 0
    errors = []
    for index, row in enumerate(rows_from_upload(uploaded_file), start=2):
        if first_available(row, fuzzy_patterns=(("full name",), ("email address",))):
            was_created, error = import_sarp_candidate(row, index)
            if error:
                errors.append(error)
                continue
            created += int(was_created)
            updated += int(not was_created)
            continue

        email = cell(row, "email", "candidate email")
        first_name = cell(row, "first_name", "first name", "firstname")
        last_name = cell(row, "last_name", "last name", "lastname")
        if not email or not first_name or not last_name:
            errors.append(f"Row {index}: first_name, last_name, and email are required.")
            continue
        _, was_created = Candidate.objects.update_or_create(
            email=email,
            defaults={
                "first_name": first_name,
                "last_name": last_name,
                "phone": cell(row, "phone"),
                "location": cell(row, "location"),
                "willing_remote": bool_cell(row, "willing_remote", "remote"),
                "work_authorization": cell(row, "work_authorization", "authorization"),
                "years_experience": decimal_cell(row, "years_experience", "experience", default="0"),
                "salary_expectation": int_cell(row, "salary_expectation", "salary"),
                "skills": cell(row, "skills"),
                "education": cell(row, "education"),
                "certifications": cell(row, "certifications", "certs"),
                "summary": cell(row, "summary"),
                "relevant_experience": cell(row, "relevant_experience", "experience_summary"),
                "status": valid_candidate_status(cell(row, "status", default=Candidate.Status.ACTIVE)),
            },
        )
        created += int(was_created)
        updated += int(not was_created)
    return {"created": created, "updated": updated, "errors": errors}


def import_jobs(uploaded_file):
    created = updated = 0
    errors = []
    for index, row in enumerate(rows_from_upload(uploaded_file), start=2):
        employer_name = cell(row, "employer", "employer_name", "company")
        title = cell(row, "title", "job_title")
        if not employer_name or not title:
            errors.append(f"Row {index}: employer and title are required.")
            continue
        employer, _ = Employer.objects.get_or_create(
            name=employer_name,
            defaults={
                "contact_name": cell(row, "contact_name"),
                "contact_email": cell(row, "employer_email", "contact_email"),
                "website": cell(row, "website"),
            },
        )
        _, was_created = Job.objects.update_or_create(
            employer=employer,
            title=title,
            defaults={
                "contact_email": cell(row, "job_email", "application_email", "contact_email"),
                "location": cell(row, "location"),
                "remote_allowed": bool_cell(row, "remote_allowed", "remote"),
                "min_years_experience": decimal_cell(row, "min_years_experience", "experience", default="0"),
                "salary_min": int_cell(row, "salary_min"),
                "salary_max": int_cell(row, "salary_max"),
                "required_skills": cell(row, "required_skills", "must_have_skills"),
                "preferred_skills": cell(row, "preferred_skills", "nice_to_have_skills"),
                "work_authorization": cell(row, "work_authorization", "authorization"),
                "education_requirements": cell(row, "education_requirements", "education"),
                "certification_requirements": cell(row, "certification_requirements", "certifications"),
                "description": cell(row, "description"),
                "must_have_notes": cell(row, "must_have_notes"),
                "status": cell(row, "status", default=Job.Status.OPEN),
            },
        )
        created += int(was_created)
        updated += int(not was_created)
    return {"created": created, "updated": updated, "errors": errors}
